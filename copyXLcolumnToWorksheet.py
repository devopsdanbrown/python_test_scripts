import openpyxl

# Open the Excel file
workbook = openpyxl.load_workbook('your_file.xlsx')

# Get the first sheet
sheet = workbook.active

# Create a new workbook
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Read the values from the first column and copy to the new worksheet
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

# Save the new workbook
new_workbook.save('new_file.xlsx')